from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

wb = load_workbook('balances.xlsx')
ws = wb.active
ws.title = "New Title"
# 设置行高和列宽
ws.column_dimensions["D"].width = 17.2
for i in range(2, ws.max_row + 1):
    ws.row_dimensions[i].height = 131
# 居中对齐
align = Alignment(horizontal='center', vertical='center')
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        ws.cell(i, j).alignment = align
# 单元格数据
ws['A4'] = 2
d = ws.cell(row=4, column=2, value=10)
# 按名称插入图片
# img1 = Image('img/111.jpg')
# ws.add_image(img1, 'D2')
# img1.width, img1.height = (125, 175)
img_list = []
for row in ws.iter_rows(min_row=2):
    img_name = row[2].value
    print(img_name)
    img_list.append(img_name)
print(img_list)
k = 0
m = 2
for row in ws.iter_rows(min_row=2):
    img_name = img_list[k]
    img_path = str("img/" + img_name)    
    img = Image(img_path)
    img.width, img.height = (125, 175)
    insert_location = str("D" + str(m))
    ws.add_image(img, insert_location)
    k += 1
    m += 1
# 保存
wb.save('balances2.xlsx')
