import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By


# 定义获取评价函数
def evaluate(url):
    # 创建Chrome浏览器对象
    browser = webdriver.Firefox()
    # 加载指定的页面
    browser.get(url)
    # 设置隐式等待时间为20秒
    browser.implicitly_wait(20)
    # 尝试查找元素
    # element = browser.find_element(By.CLASS_NAME, 'teacher-dec')
    # 提取评语
    element1 = browser.find_element(By.XPATH, '/html/body/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div/span[2]')
    # 保存评语
    comment = element1.text
    # 关闭浏览器
    browser.close()
    # 返回结果
    return (comment)


# 加载一个工作簿
wb = openpyxl.load_workbook('data.xlsx')
# 获取工作表
sheet = wb.worksheets[0]
# 获取评价并写入excel
for row_ch in range(1, sheet.max_row + 1):
    url = sheet[f'I{row_ch}'].value
    # 调用函数获取评价
    comment = evaluate(url)
    # 写入
    sheet.cell(row_ch, 10, comment)
    wb.save('data.xlsx')
    print(f'已完成第{row_ch}个', end='\n')
