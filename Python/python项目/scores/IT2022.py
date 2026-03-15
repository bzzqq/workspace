from openpyxl import load_workbook

# 读取教学记录表，将学号保存到student_number
wb = load_workbook('大学信息技术 化药2201-2205.xlsx')
ws = wb.active
student_number = []
for row in ws.iter_rows(min_row=7):
    student_number.append(row[2].value)
# print(student_number)

# 读取成绩1
wb1 = load_workbook('2270689332022大学计算机（先导课）_59945893_本部.xlsx')
ws1 = wb1.active
student_score1 = {}
for row in ws1.iter_rows(min_row=4):
    student_score1[int(row[1].value)] = float(row[9].value)
# print(student_score1)

for number, score in student_score1:
    for row in ws.iter_rows(min_row=7):
        if row[2].value == number:
            row[7].value = score
wb.save('balances2.xlsx')